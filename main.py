from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from bs4 import BeautifulSoup
from time import sleep
import openpyxl as xl
import os
# from selenium.common.exceptions import TimeoutException
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.keys import Keys

def printTerminal(name, price, url):
    print(f'{i - 1}. {name} \n {price} \n {url}')

# Init
edited_excel_path = "D:\\PythonProjects\\TopedCrawler\\products_added.xlsx"
if os.path.exists(edited_excel_path):
    os.remove(edited_excel_path)

PATH = "C:\\Program Files (x86)\\chromedriver.exe"
driver = webdriver.Chrome(PATH)
wb = xl.load_workbook('products.xlsx')
sheet = wb['Sheet1']


print('What are you looking for?')
product_name = input()
print('Filter certain words [seperate with comma (,)]:')
filter_words = input()
filter_words = filter_words.lower().split(',')
product_name = product_name.split()

url_toped = "https://www.tokopedia.com/search?st=product&q="
url_shopee = "https://shopee.co.id/search?keyword="
url_bukalapak = "https://www.bukalapak.com/products?search%5Bkeywords%5D="

for name in product_name:
    url_toped += name + '%20'
    url_shopee += name + '%20'
    url_bukalapak += name + '%20'


# Tokopedia
driver.get(url_toped)
print("Getting products from Tokopedia...")
soup = BeautifulSoup(driver.page_source, features="lxml")
products = soup.find_all('div', class_='css-1g20a2m')

i = 2
for product in products:
    try:
        productName = product.find('div', class_='css-18c4yhp').text
        productPrice = product.find('div', class_='css-rhd610').text
        productPrice = productPrice[2:]
        productURL = product.find('div', class_='css-1ehqh5q').a.get('href')

        if any(x in productName.lower() for x in filter_words):
            pass
        else:
            printTerminal(productName, productPrice, productURL)
            sheet.cell(i, 1).value = productName
            sheet.cell(i, 2).value = productPrice
            sheet.cell(i, 3).value = productURL

    except AttributeError:
        i -= 1

    i += 1


# Shopee
chrome_options = Options()
chrome_options.add_argument('disable-notifications')
chrome_options.add_argument('--disable-infobars')
chrome_options.add_argument('start-maximized')
chrome_options.add_argument('user-data-dir=C:\\Users\\username\\AppData\\Local\\Google\\Chrome\\User Data\\Default')
chrome_options.add_argument("disable-infobars")
chrome_options.add_experimental_option("prefs", {
    "profile.default_content_setting_values.notifications": 2
})

driver.get(url_shopee)
print("Getting products from Shopee...")
WebDriverWait(driver, 5)
sleep(2)

soup = BeautifulSoup(driver.page_source, features="lxml")
products = soup.find_all('div', class_='col-xs-2-4 shopee-search-item-result__item')

for product in products:
    try:
        productName = product.find('div', class_='_1NoI8_ _16BAGk').text
        productPrice = product.find('span', class_='_341bF0').text
        productURL = product.find('a', {'data-sqe': 'link'}).get('href')
        productURL = 'https://shopee.co.id/' + productURL

        if any(x in productName.lower() for x in filter_words):
            pass
        else:
            printTerminal(productName, productPrice, productURL)
            sheet.cell(i, 1).value = productName
            sheet.cell(i, 2).value = productPrice
            sheet.cell(i, 3).value = productURL

    except AttributeError:
        i -= 1

    i += 1


# Bukalapak
driver.get(url_bukalapak)
print("Getting products from Bukalapak...")
soup = BeautifulSoup(driver.page_source, features="lxml")
products = soup.find_all('div', class_='bl-flex-item mb-8')

for product in products:
    try:
        productName = product.find('a', class_='bl-link').text.strip()
        productPrice = product.find('p', class_='bl-text bl-text--subheading-2 bl-text--semi-bold bl-text--ellipsis__1').text.strip()
        productPrice = productPrice[2:]
        productURL = product.find('a', class_='bl-link').get('href')

        if any(x in productName.lower() for x in filter_words):
            pass
        else:
            printTerminal(productName, productPrice, productURL)
            sheet.cell(i, 1).value = productName
            sheet.cell(i, 2).value = productPrice
            sheet.cell(i, 3).value = productURL

    except AttributeError:
        i -= 1

    i += 1

wb.save("products_added.xlsx")
driver.close()
